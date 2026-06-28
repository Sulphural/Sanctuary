from openpyxl import load_workbook
import sys

# Read the Excel file
excel_file = 'Packets.xlsx'

try:
    # Load the workbook
    wb = load_workbook(excel_file)
    print(f"Sheets in the workbook: {wb.sheetnames}\n")
    
    # Search for relevant terms
    search_terms = ['consumable', 'transform', 'camera', 'mount', 'pet', 'appearance', 'model', 'composite']
    
    # Read each sheet
    for sheet_name in wb.sheetnames:
        print(f"\n{'='*80}")
        print(f"Sheet: {sheet_name}")
        print('='*80)
        
        ws = wb[sheet_name]
        
        # Get headers from first row
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        print(f"\nColumns: {headers}")
        print(f"Total rows: {ws.max_row - 1}\n")
        
        # Display all data
        print("All data:")
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            print(f"Row {row_idx}: {row}")
        
        # Search in all cells for relevant terms
        print(f"\n\nSearching for relevant terms in sheet '{sheet_name}':")
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            for col_idx, cell_value in enumerate(row):
                if cell_value:
                    cell_str = str(cell_value).lower()
                    for term in search_terms:
                        if term in cell_str:
                            col_name = headers[col_idx] if col_idx < len(headers) else f"Column {col_idx}"
                            print(f"  Row {row_idx}, {col_name}: {cell_value}")
                            break
        
except Exception as e:
    print(f"Error reading Excel file: {e}")
    import traceback
    traceback.print_exc()
